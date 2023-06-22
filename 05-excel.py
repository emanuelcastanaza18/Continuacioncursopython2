import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

#Esto devuelve un listado con los nombres que contiene el archivo
# print(wb.sheetnames)
# #Si quisieramos obtener una hoja en particular
# print(wb['Hoja1'])
# #Para indicar la hoja de calculo que se encuentra activa
# hoja = wb.active
# print(hoja)
# #Para crear una nueva hoja
# wb.create_sheet('Hoja2')
# print(wb.sheetnames)

# #para seleccionar una hoja y cambiarle el nombre
# hoja = wb['Hoja2']
# hoja.title = 'Hoja2'

#Cantidad de columnas y filas
# print(
#     hoja.max_row,
#     hoja.max_column
# )

#Para seleccionar una celda en particular
# celda = hoja['A1']
# print(celda.value)

#Para cambiar el valor de la celda 
# celda = hoja['A1']
# celda.value = 'Hola Mundo'
# print(celda.value)

#Otra forma seria
# celda = hoja['A1']
# celda.value = 'Hola Mundo'

# celda2 = hoja.cell(row=3, column=2)
# print(celda2.value,
#       celda2.row,
#       celda2.column,
#       celda2.coordinate
# )

#para acceder a todos los valores con un for
# celda = hoja['A1']
# celda.value = 'Hola Mundo'

# celda2 = hoja.cell(row=3, column=2)

# for fila in range(1, hoja.max_row + 1):
#     for columna in range(1, hoja.max_column + 1):
#         celda = hoja.cell(row=fila, column=columna)
#         print(fila, columna, celda.value)

#Como obtener una fila o columna en particular
# celda = hoja['A1']
# celda.value = 'Hola Mundo'

# celda2 = hoja.cell(row=3, column=2)

# for fila in range(1, hoja.max_row + 1):
#     for columna in range(1, hoja.max_column + 1):
#         celda = hoja.cell(row=fila, column=columna)

# columna = hoja['A']
# fila = hoja['1']
# print(columna)
# print(fila)

#Para agregar mas info a las filas

# celda = hoja['A1']
# celda.value = 'Hola Mundo'

# celda2 = hoja.cell(row=3, column=2)

# for fila in range(1, hoja.max_row + 1):
#     for columna in range(1, hoja.max_column + 1):
#         celda = hoja.cell(row=fila, column=columna)

# columna = hoja['A']
# fila = hoja['1']

# hoja.append([1, 2, 3])
# print(hoja.rows)

#Para eliminar filas o columnas

# celda = hoja['A1']
# celda.value = 'Hola Mundo'

# celda2 = hoja.cell(row=3, column=2)

# for fila in range(1, hoja.max_row + 1):
#     for columna in range(1, hoja.max_column + 1):
#         celda = hoja.cell(row=fila, column=columna)

# columna = hoja['A']
# fila = hoja['1']

# hoja.append([1, 2, 3])
# print(hoja.rows)
# hoja.delete_cols(1, 1)
# hoja.delete_rows(1,1)

#Crearemos un nuevo archivo excel para ver los datos que se eliminaron
celda = hoja['A1']
celda.value = 'Hola Mundo'

celda2 = hoja.cell(row=3, column=2)

for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)

columna = hoja['A']
fila = hoja['1']

hoja.append([1, 2, 3])
print(hoja.rows)
hoja.delete_cols(1, 1)
hoja.delete_rows(1,1)
wb.save('nuevo.xlsx')